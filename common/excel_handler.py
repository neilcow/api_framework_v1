import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class ExcelHandler(object):
    def __init__(self, file):
        self.file = file

    def open_sheet(self, name) -> Worksheet:
        """
            在方法后面加 -> 指向类型，代表这个方法返回的类型是这个
            加了Worksheet 类，下面的sheet就会有提示
        :param name:
        :return:
        """
        wb = openpyxl.load_workbook(self.file)
        sheet_name = wb[name]
        return sheet_name

    # 获取表头数据
    def header(self, sheet_name):
        sheet = self.open_sheet(sheet_name)
        headers = []
        for i in sheet[1]:
            headers.append(i.value)
        return headers

    # 获取所有数据
    def read(self, sheet_name):
        sheet = self.open_sheet(sheet_name)
        rows = list(sheet.rows)[1:]

        data = []
        for row in rows:
            data_row = []
            for cell in row:
                data_row.append(cell.value)
                # 用字典的形式存储
                # 列表转成字典，要和header 去zip
                data_dict = dict(zip(self.header(sheet_name), data_row))
            data.append(data_dict)
        return data

    # 写入数据
    @staticmethod
    def write(file, sheet_name, row, column, new_data):
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheet_name]
        sheet.cell(row, column).value = new_data
        wb.save(file)
        wb.close()


if __name__ == '__main__':
    excel = ExcelHandler(r'd:\cases.xlsx')
    data = excel.read('login')
    print(data)
    # excel.write(r'd:\cases.xlsx', 'Sheet1', 3, 1, 'data_new')
    # print(data)